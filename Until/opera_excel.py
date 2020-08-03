import os
import xlrd
from openpyxl import load_workbook

class OperaExcel(object):

    def __init__(self, excel_path=None):
        base_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
        if excel_path == None:
            self.excel_path = base_path + "/Case/test_case.xlsx"
        self.open_excel = xlrd.open_workbook(self.excel_path)

    def _get_sheet_data(self, index=None):
        if index == None:
            index = 0
        elif not isinstance(index, int):
            raise ValueError('index must be ini')
        data = self.open_excel.sheet_by_index(index)
        return data

    def get_rows(self):
        # 获取有效总行数
        nrows = self._get_sheet_data().nrows
        return nrows

    def get_cols(self):
        # 获取有效总列数
        ncols = self._get_sheet_data().ncols
        return ncols

    def get_row_value(self, row):
        # 获取每一行的内容
        if not isinstance(row, int):
            raise ValueError('row must be int')
        data = self._get_sheet_data().row_values(row)
        return data

    def get_col_value(self, col=None):
        # 获取每一列的内容
        if col == None:
            col = 0
        data = self._get_sheet_data().col_values(col)
        return data

    def get_data_row(self, case_id):
        # 根据case_id，找到行号
        row_number = 0
        cols_value = self.get_col_value(col=None)
        for i in cols_value:
            if case_id == i:
                return row_number
            row_number += 1

    def get_excel_value(self):
        # 获取Excel中所有用例
        nrows = self.get_rows()
        values_line = []
        for i in range(1, nrows):
            value = self.get_row_value(i)
            values_line.append(value)
        return values_line

    def write_actual_res_result(self, row, actual_message, res, excel_path=None):
        if excel_path == None:
            excel_path = self.excel_path
        # 写入实际结果
        actual_col = 11
        res_col = 12
        # 生成一个已存在的对象
        wb = load_workbook(excel_path)
        # 激活shell
        active_sheet = wb.active
        # 往shell中写入数据
        active_sheet.cell(row + 1, actual_col + 1, actual_message)
        active_sheet.cell(row + 1, res_col + 1, res)
        # 保存结果
        wb.save(excel_path)

if __name__ == '__main__':
    test = OperaExcel()
    print(test.write_actual_res_result(1, '测试', 'res'))